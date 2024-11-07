import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from image_process_and_prediction import image_process, files_operations


def create_excel_file(filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # define column names
    columns = [
        "Serial", "File Name", "Label", "Sharpness Alpha", "Sharpness image_process_and_prediction", "Sharpness Gamma",  # editor data
        "Gamma Correction", "Filter Strength", "Template Window", "Search Window", "Clip Limit",  # editor data
        "Sharpness", "Contrast", "Brightness",
        "Average Color Intensity", "Color Standard Deviation", "Number of Edges",
        "Average Edge Length", "Edge Orientation (Horizontal)", "Edge Orientation (Vertical)",
        "Entropy", "Histogram Mean", "Histogram Standard Deviation",
        "Texture Features (Local Binary Patterns)", "Frequency Mean",
        "Frequency Standard Deviation", "Foreground to Background Ratio",
        "Dominant Color", "Gradient Magnitude", "Gradient Orientation",
        "Peak Signal-to-Noise Ratio (PSNR)", "Structural Similarity Index (SSIM)",
        "Clarity Score", "Smoothness"
    ]

    # write column names to the first row and make them bold
    for col_num, column_title in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    workbook.save(filename)

def update_excel_data(filename, data, index, alpha, beta, gamma, label, gamma_correction,
                      filter_strength, template_window, search_window, clip_limit):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        values = list(data.values())
        for i, value in enumerate(values):
            if isinstance(value, list):
                values[i] = str(value)
        sheet.append(values)

        # Updating existing data at a specific row
        row = index
        if not sheet.cell(row=row, column=1).value:
            sheet.cell(row=row, column=1, value=row - 1)

        if label is not None and not sheet.cell(row=row, column=3).value:
            sheet.cell(row=row, column=3, value=label)

        if alpha is not None and not sheet.cell(row=row, column=4).value:
            sheet.cell(row=row, column=4, value=alpha)
        if beta is not None and not sheet.cell(row=row, column=5).value:
            sheet.cell(row=row, column=5, value=beta)
        if gamma is not None and not sheet.cell(row=row, column=6).value:
            sheet.cell(row=row, column=6, value=gamma)

        if gamma_correction is not None and not sheet.cell(row=row, column=7).value:
            sheet.cell(row=row, column=7, value=gamma_correction)

        if filter_strength is not None and not sheet.cell(row=row, column=8).value:
            sheet.cell(row=row, column=8, value=filter_strength)
        if template_window is not None and not sheet.cell(row=row, column=9).value:
            sheet.cell(row=row, column=9, value=template_window)
        if search_window is not None and not sheet.cell(row=row, column=10).value:
            sheet.cell(row=row, column=10, value=search_window)
        if clip_limit is not None and not sheet.cell(row=row, column=11).value:
            sheet.cell(row=row, column=11, value=clip_limit)

        workbook.save(filename)
    except Exception as e:
        print(f"Error reading or updating Excel file: {e}")
        return None

def write_current_json_to_excel():

    image_path = files_operations.get_path("IMAGE_PATH")
    index = files_operations.read_from_txt(files_operations.get_path("INDEX_FILE"))
    features = image_process.process_image(image_path)
    (alpha, beta, gamma, label, gamma_correction, filter_strength, template_window, search_window,
     clip_limit) = files_operations.read_json_data()

    update_excel_data(files_operations.get_path("EXCEL_PATH"), features, index, alpha, beta, gamma, label, gamma_correction,
                      filter_strength, template_window, search_window, clip_limit)
    files_operations.write_to_txt(index + 1, files_operations.get_path("INDEX_FILE"))

def style_excel(filename, index):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        # define the cell style
        calibri_font = Font(name='Calibri', size=11)
        centered_alignment = Alignment(horizontal='center', vertical='center')
        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        # apply the style to all cells in the worksheet
        for row in sheet.iter_rows(min_row=1, max_row=index, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = calibri_font
                cell.alignment = centered_alignment
                cell.border = border_style

        # save the workbook after styling
        workbook.save(filename)
        print("Excel file styled successfully.")
    except Exception as e:
        print(f"An error occurred while styling Excel file: {e}")